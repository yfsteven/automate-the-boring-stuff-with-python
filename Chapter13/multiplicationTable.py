#! python3
# multiplicationTable - a program that creates a nxn multiplication table on excel
import openpyxl, sys

wb = openpyxl.Workbook()
sheet = wb.active

if len(sys.argv) == 2:
    n = int(sys.argv[1])
    for i in range(1, n+1):
        sheet['A' + str(i+1)] = i
        sheet.cell(row=1, column=i+1).value = i
    row_max = sheet.max_row
    column_max = sheet.max_column
    for k in range(row_max-1):
        for j in range(column_max-1):
            sheet.cell(row=k+2, column=j+2).value = sheet.cell(row=1, column=j+2).value * sheet.cell(row=k+2, column=1).value
    wb.save('multiplicationTable.xlsx')
