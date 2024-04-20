import openpyxl, datetime


wbook = input('Enter workbook\n')
wb = openpyxl.load_workbook(wbook)


sheet = wb.active
rowMax = sheet.max_row
columnMax = sheet.max_column


sheetData2 = {}


new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

for i in range(columnMax): #Inverse the original spreadsheet inside sheet data
    for j in range(rowMax):
        sheetData2.setdefault(str(j+1), {str(i+1):''})
        if type(sheet.cell(row=j+1,column=i+1).value) is datetime.datetime:
            sheetData2[str(j+1)][str(i+1)] = sheet.cell(row=j+1,column=i+1).value.strftime('%x %X')
        else:
            sheetData2[str(j+1)][str(i+1)] = sheet.cell(row=j+1,column=i+1).value

for col in sheetData2: #Add sheet data into a new workbook
    for i in range(columnMax):
        new_sheet.cell(row=i+1,column=int(col)).value = sheetData2[str(col)][str(i+1)]
new_workbook.save('updated'+wbook)
