#! python3
# convertExcelToCSV.py - gather all excel file and convert them into CSV file
import openpyxl, csv, os

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue
    try:
        wb = openpyxl.load_workbook(excelFile)
        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]
            csvFile = open(excelFile.strip('.xlsx') + sheetName +'.csv', 'w', newline='')
            csvWriter = csv.writer(csvFile)
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                csvWriter.writerow(rowData)
            csvFile.close()
    except Exception as err:
        print(excelFile + ': '+ str(err))


